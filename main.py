# -------------------------------------------------------------------------------
# Name:        process_survey123_datafile
# Purpose:
#
# Author:      Adrian Kitchingman & Alexander Hopper.
#
# Created:     13/05/2022
# Copyright:   (c) ak34 2022
# Licence:     <your licence>
# Last Update: 06/02/2024

# ~~~~ TO DO ~~~~~~~~~~~~
#   Add timestamp
#   Add name of input file
#   Error Check output file not open
#   Navigate to input file
#       https://stackoverflow.com/questions/9319317/quick-and-easy-file-dialog-in-python
# -------------------------------------------------------------------------------
#

# ==========================================================================================================================================#
# ==========================================================Preference Changes:=============================================================#
# ==========================================================================================================================================#

##export_default = False


##input_type = 'Fish_Survey_v1'
input_type = 'Fish_Survey_v2'
##input_type = ''

# The following is to order how each page is presented in the results.

# The data column will not be output if the index is == -1.
# The data column will move to the corresponding index value, e.g. if 'site_code' is at index 1,
# then [-1, 0, -1, -1, ...] will only output 'site_code' and place it at the start.
# A 'j' will join that column and the following column, and place it in the position of the value after the 'j'.
# A list of [0, 1, 2, 3, ...] will not change any order.
# This will also not take into account 'ObjectID', i.e. 'GlobalID' is the first element.
if input_type == 'Fish_Survey_v1':
    survey_template = [-1, 1, 4, 5, 'j', 6, 7, 8, 9, 10, 11, 12, 0, 13, -1, 14, 15, 16, 17, 18, 19, 20, 21, -1, -1, -1,
                       -1, -1, 2, 3]
    location_template = [-1, -1, -1, -1, -1, -1, -1, 0, 1, 2,
                         3]  # Keep in mind [... x, y] will become ... x_start, y_start, x_end, y_end]
    shot_template = [-1, 0, 1, 2, 3, 4, 5, 6, 7, 8, -1, -1, -1, -1, -1]
    obs_template = [-1, -1, -1, 0, 1, 2, -1, -1, -1, -1, -1, -1]
    sample_template = [-1, -1, -1, 0, 1, 2, 3, 4, 5, 6, 7, -1, 8, 9, 10, -1, -1, -1, -1, -1]

elif input_type == 'Fish_Survey_v2':

    survey_template = [-1, 1, 4, 5, 'j', 6, 7, 8, 9, 10, 11, 12, 0, 13, -1, 14, 15, 16, 17, 18, 19, 20, 21, -1, -1,
                       -1, -1, 2, 3]
    location_template = [-1, -1, -1, -1, -1, -1, -1, 0, 1, 2,
                         3]  # Keep in mind [... x, y] will become ... x_start, y_start, x_end, y_end]
    shot_template = [-1, 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, -1, -1, -1]
    obs_template = [-1, -1, -1, 0, 1, 2, -1, -1, -1, -1, -1, -1]
    sample_template = [-1, -1, -1, 0, 1, 2, 3, 4, 5, 6, 7, -1, 8, 9, 10, -1, -1, -1, -1, -1]

else:

    survey_template = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26,
                       27, 28, 29]
    location_template = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9,
                         10]  # Keep in mind [... x, y] will become ... x_start, y_start, x_end, y_end]
    shot_template = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]
    obs_template = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]
    sample_template = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]


# Names of columns to sort for raw and tally sheets:
raw_sorters = ['Site_GlobalID', 'survey_date', 'section_number', 'species', 'observed', 'section_collected']
tally_sorters = ['Site_ID', 'Section_Number', 'Species']

# ==========================================================================================================================================#
# ========================================================End Preference Changes:===========================================================#
# ==========================================================================================================================================#


# import io
import os
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
import process_survey123_field_data_classes as cls
import process_survey123_field_data_functions as func
import local_vars as localvars #### local_vars.py is py file holding local paths (is git ignored)
from tkinter import *
import tkinter.messagebox
from tkinter.filedialog import askopenfilename  # as fd
import random
import re

root = Tk()
root.withdraw()
root.update()

io_path = localvars.io_path
filename = askopenfilename(initialdir=io_path, title="Open Survey123 XLSX File")
in_xlfile = os.path.basename(filename)
io_path = filename.replace(in_xlfile, '')


if (len(filename) == 0):
    print('*** PROCESSING CANCELLED')
    pass
else:

    while True:  # repeat until the try statement succeeds
        try:

            workbook = load_workbook(filename)
            break

        except IOError:

            answer = tkinter.messagebox.askokcancel("Open File Error",
                                                    "Could not open file! Please close Excel. Press OK to retry.")
    root.destroy()

    print('*** OUTPUT FORMAT: {0}'.format(input_type))

    print(workbook.sheetnames)

    # # ## READ SITE SURVEY DATA =======================================================================================

    survey_list = func.read_in_excel_tab(workbook.worksheets[0])
    survey_list_header = list(func.read_in_excel_tab_header(workbook.worksheets[0]))

    # # ## READ location DATA =======================================================================================

    loc_list = func.read_in_excel_tab(workbook['site_location_repeat_1'])
    loc_list_header = list(func.read_in_excel_tab_header(workbook['site_location_repeat_1']))

    # # ## READ SHOT DATA =======================================================================================

    shot_list = func.read_in_excel_tab(workbook['shot_repeat_2'])
    shot_list_header = list(func.read_in_excel_tab_header(workbook['shot_repeat_2']))

    # # ## READ observed DATA =======================================================================================

    obs_list = func.read_in_excel_tab(workbook['observed_fish_repeat_3'])
    obs_list_header = list(func.read_in_excel_tab_header(workbook['observed_fish_repeat_3']))

    # # ## READ sampled DATA =======================================================================================

    sample_list = func.read_in_excel_tab(workbook['fish_sample_repeat_4'])
    sample_list_header = list(func.read_in_excel_tab_header(workbook['fish_sample_repeat_4']))

    # Sort the samples so any defined shots are at the top.
    sample_list.sort(key=lambda x: 0 if x[sample_list_header.index('section_number_samp')] is None else int(
        x[sample_list_header.index('section_number_samp')]), reverse=True)
    sample_list.sort(key=lambda x: x[sample_list_header.index('ParentGlobalID')])

    # Change x,y to x_start,y_start,x_end,y_end in location header.
    loc_list_header[loc_list_header.index('x')] = 'x_coordinate'
    loc_list_header[loc_list_header.index('y')] = 'y_coordinate'
    loc_list_header.append('finish_x_coordinate')
    loc_list_header.append('finish_y_coordinate')

    print('Gathering Data...')
    raw_data = []
    tally_results = []

    ## ----------------------------------------------------------------------------------------------------------------------
    ## ----------------------------------------------------------------------------------------------------------------------
    ## ----- LOOP THROUGH Surveys (site), Loc, Shots and Obs WORKSHEETS to build raw_data base and tally list ---------------
    ## ----------------------------------------------------------------------------------------------------------------------
    ## ----------------------------------------------------------------------------------------------------------------------

    for svy in survey_list:
        survey_list_current = list(svy)
        site_id = survey_list_current[survey_list_header.index('GlobalID')]

        # Change gear_type name and set section_condition to (UN)FISHABLE
        survey_list_current[survey_list_header.index('gear_type')] = func.gear_types[
            survey_list_current[survey_list_header.index('gear_type')]]
        if survey_list_current[survey_list_header.index('section_condition')].lower() == 'yes':
            survey_list_current[survey_list_header.index('section_condition')] = 'FISHABLE'
        else:
            survey_list_current[survey_list_header.index('section_condition')] = 'UNFISHABLE'
            print('*** CAUTION: Survey marked UNFISHABLE for Site ID: {0}'.format(site_id))

        creator = survey_list_current[survey_list_header.index('Creator')]

        # Loop through locations survey sheet - filtering for each survey.
        for lcs in filter(lambda x: x[loc_list_header.index('ParentGlobalID')] == survey_list_current[
            survey_list_header.index('GlobalID')], loc_list):
            loc_list_current = list(lcs)

            # Only add data if starting coordinate to not double up on data.
            if loc_list_current[loc_list_header.index('point_location')] == 'site_start':

                # See if there is another location entry with same GlobalID for end coordinates.
                pair = False
                x_finish = 0
                y_finish = 0
                for same_location in filter(lambda x: x[loc_list_header.index('GlobalID')] == loc_list_current[
                    loc_list_header.index('GlobalID')], loc_list):
                    if same_location[loc_list_header.index('point_location')] == 'site_end':
                        pair = True
                        x_finish = same_location[loc_list_header.index('x_coordinate')]
                        y_finish = same_location[loc_list_header.index('y_coordinate')]

                loc_list_current.append(x_finish)
                loc_list_current.append(y_finish)

                # Loop through shots sheet - filtering for location.
                sts = list(filter(lambda x: x[shot_list_header.index('ParentGlobalID')] == loc_list_current[
                    loc_list_header.index('ParentGlobalID')], shot_list))

                #If shots have been recorded
                if len(sts) > 0:

                    for shot_list_current in sts:
                        shot_list_current = list(shot_list_current)

                        section_index = shot_list_header.index('section_number')
                        if shot_list_current[section_index] is None and len(sts) == 1:
                            shot_list_current[section_index] = 1

                        shot_id = shot_list_current[shot_list_header.index('GlobalID')]
                        section_number = shot_list_current[section_index]

                        # Create filler for samples:
                        sample_list_current = [None] * len(sample_list_header)

                        # Loop through observations - filtering for shots. And removing shots with 'obs_ts' == None.
                        obs = list(filter(lambda x: x[obs_list_header.index('ParentGlobalID')] == shot_list_current[
                            shot_list_header.index('GlobalID')] and x[obs_list_header.index('obs_ts')] is not None,
                                          obs_list))

                        #If Observations are recorded
                        if len(obs) > 0:
                            for obs_list_current in obs:
                                obs_list_current = list(obs_list_current)
                                obs_id = obs_list_current[obs_list_header.index('GlobalID')]

                                s_custom = obs_list_header.index('species_obs_custom')
                                s_new = obs_list_header.index('species_new')
                                s_obs = obs_list_header.index('species_obs')
##                                if survey_list_current[survey_list_header.index('section_condition')] == 'FISHABLE':
                                # Build object for output
                                # Find ID Indices:
                                ID_Indices = [survey_list_header.index('GlobalID'),
                                              loc_list_header.index('GlobalID'),
                                              shot_list_header.index('GlobalID'),
                                              obs_list_header.index('GlobalID'),
                                              sample_list_header.index('GlobalID'), ]
                                raw_data.append(cls.resultObject(survey_list_current,
                                                                 loc_list_current,
                                                                 shot_list_current,
                                                                 obs_list_current,
                                                                 sample_list_current,
                                                                 creator,
                                                                 ID_Indices))

                                if obs_list_current[obs_list_header.index('section_collected')] is None:
                                    obs_list_current[obs_list_header.index('section_collected')] = 0

                                # Determine species from columns custom, new and obs.
                                if obs_list_current[s_custom] is not None or obs_list_current[s_new] is not None or \
                                        obs_list_current[s_obs] is not None:
                                    if obs_list_current[s_custom] is None:
                                        species = obs_list_current[s_obs]
                                    elif obs_list_current[s_custom] is not None:
                                        species = obs_list_current[s_custom]
                                    else:
                                        species = obs_list_current[s_new]

                                    # Enter correct species
                                    obs_list_current[s_obs] = species

                                    collected = obs_list_current[obs_list_header.index('section_collected')]
                                    observed = obs_list_current[obs_list_header.index('observed')]

                                    if collected is None:
                                        obs_list_current[obs_list_header.index('section_collected')] = 0
                                    if observed is None:
                                        obs_list_current[obs_list_header.index('observed')] = 0

                                    if collected != 0 or observed != 0 or species == 'No Fish':
                                        # Build object for tally results:
                                        # site_id already defined (used to be here
                                        # section_number already defined (used to be here)
                                        # Species already defined
                                        # Collected already defined
                                        # Observed already defined
                                        # shot_id  already defined (used to be here)
                                        #  obs_id  already defined (used to be here)
                                        # Creator already defined

                                        # Send tally data for output
                                        tally_results.append(
                                            [site_id, section_number, species, collected, observed, collected, shot_id,
                                             obs_id, creator])
                                    else:
                                        print('Notice: SKIPPED Observation - No collected or observed record for {0}. Obs GlobalID: {1}'.format(species, obs_id))


                        else:
                            #If no Observations are recorded add a 'no fish' obs (if no samples are recorded)
##                            if survey_list_current[survey_list_header.index('section_condition')] == 'FISHABLE':

                            smpls = list(filter(lambda x: x[sample_list_header.index('ParentGlobalID')] == site_id and x[sample_list_header.index('species_samp')] != 'No Fish', sample_list))

                            #if there are samples for the site but the survey only has one shot then add the species' to the one shot
                            if len(smpls) > 0 and len(sts) == 1:

                                    func.add_samples_to_output_and_tally(smpls, sample_list_header, section_number, shot_list_current, loc_list_current, survey_list_current,
                                        raw_data, survey_list_header, loc_list_header, shot_list_header, obs_list_header, tally_results)

                            else: #If there are no samples recorded for this shot then add a No Fish observations

                                print('Notice: Created NO FISH shot for site {0} : shot No {1} '.format(survey_list_current[survey_list_header.index('GlobalID')], shot_list_current[section_index]))
                                obs_list_current = [None] * len(obs_list_header)
                                obs_list_current[obs_list_header.index('section_collected')] = 0
                                obs_list_current[obs_list_header.index('species_obs')] = 'No Fish'
                                sample_list_current[sample_list_header.index('species_samp')] = 'No Fish'
                                # Build object for output
                                # Find ID Indices:
                                ID_Indices = [survey_list_header.index('GlobalID'),
                                              loc_list_header.index('GlobalID'),
                                              shot_list_header.index('GlobalID'),
                                              obs_list_header.index('GlobalID'),
                                              sample_list_header.index('GlobalID'), ]
                                raw_data.append(cls.resultObject(survey_list_current,
                                                                 loc_list_current,
                                                                 shot_list_current,
                                                                 obs_list_current,
                                                                 sample_list_current,
                                                                 creator,
                                                                 ID_Indices))

                                # Send empty shot to tally for output
                                tally_results.append([site_id, section_number, 'No Fish', None, None, None, shot_id,
                                                 None, creator])


                # If no shots exist, add 1 shot if fishable or samples present, and add site information.
                else:

##                    if survey_list_current[survey_list_header.index('section_condition')] == 'FISHABLE':

                    shot_list_current = [None] * len(shot_list_header)
                    shot_list_current[shot_list_header.index('section_number')] = 1
                    shot_list_current[shot_list_header.index('ParentGlobalID')] = site_id

                    # Create filler for observations:
                    obs_list_current = [None] * len(obs_list_header)
                    obs_list_current[obs_list_header.index('section_collected')] = 0

                    smpls = list(filter(lambda x: x[sample_list_header.index('ParentGlobalID')] == site_id, sample_list))

                    if len(smpls) > 0:
                        func.add_samples_to_output_and_tally(smpls, sample_list_header, 1, shot_list_current, loc_list_current, survey_list_current,
                            raw_data, survey_list_header, loc_list_header, shot_list_header, obs_list_header, tally_results)

                    #If no samples are present add a single No Fish shot
                    else:
                        print('Notice: Created NO FISH shot (1) for site {0} '.format(survey_list_current[survey_list_header.index('GlobalID')]))
                        shot_list_current = [None] * len(shot_list_header)
                        shot_list_current[shot_list_header.index('section_number')] = 1
                        shot_list_current[shot_list_header.index('ParentGlobalID')] = site_id
                        obs_list_current[obs_list_header.index('section_collected')] = 0
                        obs_list_current[obs_list_header.index('species_obs')] = 'No Fish'
                        sample_list_current[sample_list_header.index('species_samp')] = 'No Fish'
                        # Build object for output
                        # Find ID Indices:
                        ID_Indices = [survey_list_header.index('GlobalID'),
                                      loc_list_header.index('GlobalID'),
                                      shot_list_header.index('GlobalID'),
                                      obs_list_header.index('GlobalID'),
                                      sample_list_header.index('GlobalID'), ]
                        raw_data.append(cls.resultObject(survey_list_current,
                                                         loc_list_current,
                                                         shot_list_current,
                                                         obs_list_current,
                                                         sample_list_current,
                                                         creator,
                                                         ID_Indices))

                        # Send empty shot to tally for output
                        tally_results.append([site_id, 1, 'No Fish', None, None, None, shot_id,
                                         None, creator])

                # Define tally header, but only once
                try:
                    tally_header
                except:
                    tally_header = ['Site_ID', 'Section_Number', 'Species', 'Collected', 'Observed',
                                    'Collected_Tally', 'shot_id', 'obs_id', 'Creator']


    ## ----------------------------------------------------------------------------------------------------------------------
    ## ----------------------------------------------------------------------------------------------------------------------
    ## ----- LOOP THROUGH SAMPLES WORKSHEET ---------------------------------------------------------------------------------
    ## ----------------------------------------------------------------------------------------------------------------------
    ## ----------------------------------------------------------------------------------------------------------------------
    print('Processing Sampled Data...')

    sample_checklist = ['section_number_samp',
                        'fork_length',
                        'total_length',
                        'weight',
                        'collected',
                        'recapture',
                        'external_tag_no',
                        'pit',
                        'genetics_label',
                        'otoliths_label',
                        'fauna_notes']

    for smp in sample_list:
        sample_list_current = list(smp)
        creator = sample_list_current[sample_list_header.index('Creator')]
        s_custom = sample_list_current[sample_list_header.index('species_samp_custom')]
        s_samp = sample_list_current[sample_list_header.index('species_samp')]

        if s_custom is not None or s_samp is not None:
            species = s_samp if s_custom is None else s_custom

            sample_list_current[sample_list_header.index('species_samp')] = species
            skip_samp = TRUE
            sample_comparison = 0

            for dummy in sample_checklist:
                if sample_list_current[sample_list_header.index(dummy)] not in [0, None, 'no']:
                    skip_samp = FALSE

            if skip_samp == FALSE:

                # Find random shot to attribute sample to with same ParentGlobalID and Species.
                site_id = sample_list_current[sample_list_header.index('ParentGlobalID')]

                # If a shot is already asigned otherwise set to 0 to get a random shot
                section_number = (0 if sample_list_current[sample_list_header.index('section_number_samp')] is None else sample_list_current[sample_list_header.index('section_number_samp')])

                if section_number == 0:
                    rand_pick = func.get_random_shot(site_id, species, raw_data, obs_list_header, shot_list_header)
                else:
                    shotlist = list(filter(lambda x: x.shots[shot_list_header.index('ParentGlobalID')] == site_id and x.shots[shot_list_header.index('section_number')] == section_number
                        and x.observations[obs_list_header.index('species_obs')] != 'No Fish',  raw_data))

                    if len(shotlist) > 0:
                        rand_pick = shotlist[0]
                    else:
                        rand_pick = False
                # If a shot is found:
                if rand_pick != False:
                    # Fix collected number:
                    section_num = (rand_pick.shots[shot_list_header.index('section_number')]
                        if section_number == 0 else section_number)
                    if sample_list_current[sample_list_header.index('collected')] is None or sample_list_current[
                        sample_list_header.index('collected')] == 0:
                        sample_list_current[sample_list_header.index('collected')] = 1

                    func.adjust_species_count(sample_list_current, raw_data,
                                              sample_list_current[sample_list_header.index('ParentGlobalID')], section_num,
                                              species, survey_list_header, obs_list_header, sample_list_header,
                                              shot_list_header, tally_results, tally_header)

                    # Build obs list:
                    obs_list_current = [None] * len(obs_list_header)
                    if sample_list_current[sample_list_header.index('collected')] is None or sample_list_current[
                        sample_list_header.index('collected')] == 0:
                        obs_list_current[obs_list_header.index('section_collected')] = 1
                    else:
                        obs_list_current[obs_list_header.index('section_collected')] = sample_list_current[
                            sample_list_header.index('collected')]
                    if obs_list_current[obs_list_header.index('observed')] is None:
                        obs_list_current[obs_list_header.index('observed')] = 0

                    # Find ID Indices:
                    ID_Indices = [survey_list_header.index('GlobalID'),
                                  loc_list_header.index('GlobalID'),
                                  shot_list_header.index('GlobalID'),
                                  obs_list_header.index('GlobalID'),
                                  sample_list_header.index('GlobalID'), ]
                    # Build object for output
                    raw_data.append(cls.resultObject(rand_pick.surveys,
                                                     rand_pick.locations,
                                                     rand_pick.shots,
                                                     obs_list_current,
                                                     sample_list_current,
                                                     creator,
                                                     ID_Indices))

                    #A No Fish would have been made if no fish recorded in shot (even though in samples). Need to remove this No Fish
                    func.remove_unrequired_no_fish(raw_data, sample_list_current[sample_list_header.index('ParentGlobalID')], section_num,
                                                survey_list_header, obs_list_header, sample_list_header, shot_list_header, tally_results, tally_header)


                # If no available shots are found:
                else:

                    # if a sample has been assigned to a shot which isn't recorded in the observed then add the species
                    if sample_list_current[sample_list_header.index('section_number_samp')] not in [0, None]:

                        section_num = sample_list_current[sample_list_header.index('section_number_samp')]
                        print('Notice: Adding {0} to shot {1} in site ID: {2}'.format(species, section_num, site_id))

                        if sample_list_current[sample_list_header.index('collected')] not in [0, None]:
                            sample_list_current[sample_list_header.index('collected')] = 1

                        shotlist = list(filter(lambda x: x.shots[shot_list_header.index('ParentGlobalID')] == site_id and x.shots[shot_list_header.index('section_number')] == section_num,  raw_data))

                        if len(shotlist) == 0:
                            print('*** ERROR: Sample not added. Missing shot listed in sample: {0}'.format(sample_list_current[sample_list_header.index('GlobalID')]))
                        else:
                            obs_list_current = [None] * len(obs_list_header)
                            obs_list_current[obs_list_header.index('section_collected')] = 1

                            ID_Indices = [survey_list_header.index('GlobalID'),
                                      loc_list_header.index('GlobalID'),
                                      shot_list_header.index('GlobalID'),
                                      obs_list_header.index('GlobalID'),
                                      sample_list_header.index('GlobalID'), ]

                            raw_data.append(cls.resultObject(shotlist[0].surveys,
                                                             shotlist[0].locations,
                                                             shotlist[0].shots,
                                                             obs_list_current,
                                                             sample_list_current,
                                                             creator,
                                                             ID_Indices))

                            tally_results.append(
                                [site_id, section_num, species, collected, observed, collected, shotlist[0].shots[shot_list_header.index('GlobalID')],
                                 None, creator])

                            #A No Fish would have been made if no fish recorded in shot (even though in samples). Need to remove this No Fish
                            func.remove_unrequired_no_fish(raw_data, sample_list_current[sample_list_header.index('ParentGlobalID')], section_num,
                                                        survey_list_header, obs_list_header, sample_list_header, shot_list_header, tally_results, tally_header)

                    else:
                        #No Shot for sample to go in (but could have been handled earlier)

                        #Check sample already in raw_data
                        if func.check_sample_in_raw_data(raw_data,sample_list_current[sample_list_header.index('GlobalID')],sample_list_header) == False:
                            print('*** ERROR: No Shot for {0} found - Sample GlobalID: {1}\n'.format(species, sample_list_current[sample_list_header.index('GlobalID')]))
                        else:
                            print('Notice: No Shot Sample of {0} has been added to output for site ID: {1}\n'.format(species, sample_list_current[sample_list_header.index('ParentGlobalID')]))

            else:
                print('Notice: REMOVED NO INFO SAMPLE - Sample GlobalID: {0}'.format(sample_list_current[sample_list_header.index('GlobalID')]))

    ## ----------------------------------------------------------------------------------------------------------------------
    ## ----------------------------------------------------------------------------------------------------------------------
    ## ----------------------------------------------------------------------------------------------------------------------
    ## ----------------------------------------------------------------------------------------------------------------------

    # Correctly format and order entries as list at top of code.
    for obj in raw_data:
        survey_list_header_edit = obj.order(obj.surveys, survey_template, survey_list_header, 'survey')
        loc_list_header_edit = obj.order(obj.locations, location_template, loc_list_header, 'location')
        shot_list_header_edit = obj.order(obj.shots, shot_template, shot_list_header, 'shot')
        obs_list_header_edit = obj.order(obj.observations, obs_template, obs_list_header, 'obs')
        sample_list_header_edit = obj.order(obj.samples, sample_template, sample_list_header, 'sample')

    print('Collating data...')
    raw_data_header = survey_list_header_edit + loc_list_header_edit + shot_list_header_edit + obs_list_header_edit + sample_list_header_edit
    # Append GlobalID List and creator:
    raw_data_header += ['Site_GlobalID', 'Loc_GlobalID', 'Shot_GlobalID', 'Obs_GlobalID', 'Sample_GlobalID', 'Creator']

    for i in raw_data:
        i.collate(raw_data_header)

    raw_data_header.pop(raw_data_header.index('species_samp'))
    raw_data_header[raw_data_header.index('species_obs')] = 'species'
    func.populate_extra_collected(raw_data, raw_data_header)

    #Only deal with net attributes for particular survey versions
    if input_type in ['Fish_Survey_v2']:
        func.correct_net_gear_type(raw_data, raw_data_header)

    #remove records that have no observed or collected data (usually inserted for tally keeping purposes)
    for i in raw_data[:]:
        if i.collation[raw_data_header.index('species')].lower() != 'no fish' and i.collation[raw_data_header.index('section_collected')] <= 0 and i.collation[raw_data_header.index('observed')] == 0 and i.collation[raw_data_header.index('collected')] is None:
            ob_id = i.collation[raw_data_header.index('Obs_GlobalID')]
            raw_data.remove(i)
            print('Notice: REMOVED Obs Record (Negative collected, Zero Observed): {0}'.format(ob_id))

    ## ----------------------------------------------------------------------------------------------------------------------
    ## ----------------------------------------------------------------------------------------------------------------------
    ## ----------------------------------------------------------------------------------------------------------------------
    ## ----------------------------------------------------------------------------------------------------------------------

    print('Processing data into new format...')

    # Open new worksheet and create Raw Data & Tally Results
    wb = openpyxl.Workbook()
    ws_write = wb.active
    ws_write.title = "Raw Data"
    ws2_write = wb.create_sheet("Tally Results", 1)

    # Write header and all data.
    # Raw Data:
    row_count = 1
    func.write_row(ws_write, row_count, 1, raw_data_header)
    for result in raw_data:
        row_count += 1
        wer_species = result.collation[raw_data_header.index('species')]
        wer_species = re.sub(r'\(.*?\) *', '', wer_species)
        wer_species = wer_species.strip()
        result.collation[raw_data_header.index('species')] = wer_species
        func.write_row(ws_write, row_count, 1, tuple(result.collation))

    # Tally Data:
    row_count = 1
    func.write_row(ws2_write, row_count, 1, tally_header)
    for result in tally_results:
        row_count += 1
        func.write_row(ws2_write, row_count, 1, result)

    # Sort both pages of output by column name at top of code:
    try:
        # Calculates indices of sort columns:
        raw_sort_indices = []
        tally_sort_indices = []

        for name in raw_sorters:
            raw_sort_indices.append(raw_data_header.index(name) + 1)
        for name in tally_sorters:
            tally_sort_indices.append(tally_header.index(name) + 1)
        func.sheet_sort_rows(ws_write, 2, 0, raw_sort_indices)
        func.sheet_sort_rows(ws2_write, 2, 0, tally_sort_indices)
    except:
        print("*** ERROR Sorting Columns, Check Sort Names")

    # Resize columns to match entry lengths:
    for column_cells in ws_write.columns:
        length = max(len(str(cell.value)) for cell in column_cells) + 3
        ws_write.column_dimensions[column_cells[0].column_letter].width = length
    for column_cells in ws2_write.columns:
        length = max(len(str(cell.value)) for cell in column_cells) + 3
        ws2_write.column_dimensions[column_cells[0].column_letter].width = length

    # Add auto filter to output sheet:
    ws_write.auto_filter.ref = ws_write.dimensions
    ws2_write.auto_filter.ref = ws2_write.dimensions

    # Determine current date and export.
    dnow = datetime.now()
    fdt = dnow.strftime("%y") + dnow.strftime("%m") + dnow.strftime("%d")

    out_xlfile = in_xlfile.replace('(', '').replace(')', '')
    out_xlfile = out_xlfile.replace(".xlsx", "_FlatFormat_" + fdt + ".xlsx")

    wb.save(io_path + out_xlfile)

    print('\nFormatting complete. New Excel file is at:\n{0}\n\n'.format(io_path + out_xlfile))

# # #loop through sampled
# xl_header = list(("project",
#                     "site_code",
#                     "x_coordinate",
#                     "y_coordinate",
#                     "finish_x_coordinate",
#                     "finish_y_coordinate",
#                     "survey_date",
#                     "gear_type",
#                     "personnel1",
#                     "personnel2",
#                     "depth_secchi",
#                     "depth_max",
#                     "depth_avg",
#                     "section_condition",
#                     "time_start",
#                     "time_end",
#                     "survey_notes",
#                     "section_number",
#                     "electro_seconds",
#                     "soak_minutes_per_unit",
#                     "section_time_start",
#                     "section_time_end",
#                     "volts",
#                     "amps",
#                     "pulses_per_second",
#                     "percent_duty_cycle",
#                     "species",
#                     "fork_length",
#                     "total_length",
#                     "weight",
#                     "collected",
#                     "observed",
#                     "recapture",
#                     "pit",
#                     "external_tag_no",
#                     "genetics_label",
#                     "otoliths_label",
#                     "fauna_notes",
#                     "water_qual_depth",
#                     "ec_25c",
#                     "water_temp",
#                     "do_mgl",
#                     "do_perc",
#                     "ph",
#                     "turbidity_ntu",
#                     "chlorophyll",
#                     "Site_GlobalID",
#                     "Shot_GlobalID",
#                     "Obs_GlobalID",
#                     "Sample_GlobalID",
#                     "data_recording_x",
#                     "data_recording_y"))

# row_count += 1

# prev_sample_site_id = ''
# shots_used = []

# samples_present = False


# for smp in sample_list:

#     # #        objectid=smp[0]
#     # #        globalid=smp[1]
#     # #        section_number_samp=smp[2]
#     # #        species_samp_custom=smp[3]
#     # #        species_samp=smp[4]
#     # #        fork_length=smp[5]
#     # #        total_length=smp[6]
#     # #        weight=smp[7]
#     # #        coll=smp[8]

#     # print(smp[0])

#     # species name or custom name required
#     if smp[3] is not None or smp[4] is not None:
#         samples_present = True
#         sample_id = smp[1]

#         section_number = 0 if smp[2] is None else smp[2]
#         # #        if smp[2] is not None:
#         # #            print(smp[2])
#         species = smp[4] if smp[3] is None else smp[3]
#         fl = '' if smp[5] is None else smp[5]
#         tl = '' if smp[6] is None else smp[6]
#         w = '' if smp[7] is None else smp[7]

#         recapture = 'no' if smp[9] is None else smp[9]
#         pit = "{0}".format(smp[11]) if smp[11] is not None else ''
#         external_tag_no = '' if smp[10] is None else smp[10]
#         genetics_label = '' if smp[13] is None else smp[13]
#         otoliths_label = 'no' if smp[14] is None else smp[14]
#         fauna_notes = '' if smp[15] is None else smp[15]

#         skip_samp = FALSE
#         if (smp[8] is None or smp[8] == 0) and section_number == 0 and fl == '' and tl == '' and w == '' and recapture == 'no' and pit == '' and external_tag_no == '' and genetics_label == '' and otoliths_label == 'no' and fauna_notes == '':
#             skip_samp = TRUE

#         if skip_samp == FALSE:

#             coll = 1 if smp[8] is None or smp[8] == 0 else smp[8]
#             sample_site_id = smp[16]

#             if sample_site_id != prev_sample_site_id:
#                 # #            print(sample_site_id ,prev_sample_site_id)
#                 if prev_sample_site_id != '':
#                     # # OUTPUT EXTRA DATA
#                     row_count = func.extra_record_output(ws_write, prev_sample_site_id, row_count)
#                 prev_sample_site_id = sample_site_id
#                 # #-----get site and shot info
#                 # #-------get species
#                 # #- -----------GET RANDOM SHOT

#             shot_i = func.get_random_shot(sample_site_id, species) if section_number == 0 else str(section_number)

#             #get site/survey for the shot selected
#             sub_site_survey_info = list(filter(lambda x: x['k_site_id'] == sample_site_id and x['k_section_number'] == shot_i, func.site_survey_info))

#             extra_shot = 0
#             if len(sub_site_survey_info) == 0:

#                 extra_shot = 1
#                 shot_i = str(section_number)
#                 sub_site_survey_info = list(filter(lambda x: x['k_site_id'] == sample_site_id and x['k_section_number'] == '1', func.site_survey_info))
#                 print('*** NO SHOT ERROR for\nsite: {0}\nsample:{1}\nshot: {2}\n*** -----------'.format(sample_site_id, sample_id, shot_i))

#             if len(sub_site_survey_info) > 1:
#                 print('*** MULTIPLE SURVEY ERROR for\nsite: {0}\nshot: {1}\n*** --------------'.format(sample_site_id, shot_i))

#             # #        shot_i = s[1]
#             if isinstance(shot_i, str):
#                 shot_i = int(shot_i)

#             # shots_used.append(shot_i)
#             func.site_section_used.append(cls.SiteSections(sample_site_id, int(shot_i)))

#             for ss_row in sub_site_survey_info:
#                 func.adjust_species_shot(sample_site_id, species, str(shot_i), coll)
#                 gear_type = ss_row.gear_type
#                 section_time_start = ss_row.section_time_start
#                 section_time_end = ss_row.section_time_end
#                 electro_seconds = ss_row.electro_seconds
#                 volts = ss_row.volts
#                 amps = ss_row.amps
#                 pulses_per_second = ss_row.pulses_per_second
#                 percent_duty_cycle = ss_row.percent_duty_cycle

#                 if extra_shot == 1:
#                     ss_row.gear_type = 'EXTRA_SHOT_IN_SAMPLES'
#                     ss_row.section_time_start = ''
#                     ss_row.section_time_end = ''
#                     ss_row.electro_seconds = 0
#                     ss_row.volts = ''
#                     ss_row.amps = ''
#                     ss_row.pulses_per_second = ''
#                     ss_row.percent_duty_cycle = ''

#                 func.write_excel_row(ws_write, row_count, ss_row, shot_i, species, fl, tl, w, coll, 0, recapture, pit, external_tag_no, genetics_label, otoliths_label, fauna_notes, '', sample_id)

#                 ss_row.gear_type = gear_type
#                 ss_row.section_time_start = section_time_start
#                 ss_row.section_time_end = section_time_end
#                 ss_row.electro_seconds = electro_seconds
#                 ss_row.volts = volts
#                 ss_row.amps = amps
#                 ss_row.pulses_per_second = pulses_per_second
#                 ss_row.percent_duty_cycle = percent_duty_cycle

#                 row_count += 1
#     else:
#         samples_present = False
#         sample_id = smp[1]
#         sample_site_id = smp[16]
#         section_number = '0' if smp[2] is None else smp[2]

#         if section_number != '0':
#             # test if the shot exisit in the shot list and add if it isn't
#             site_shot_list = list(filter(lambda x: x[0] == sample_site_id and x[1] == section_number, func.sssoc_info))
#             if len(site_shot_list) == 0:
#                 func.sssoc_info.append(cls.SiteObs(sample_site_id, section_number, '', 0, 0, 0, 'IN SAMPLE INFO', ''))
#         else:
#             print('*** SPECIES NAME ERROR IN SAMPLES: {0}'.format(sample_id))


# # #####################################################################################################################
# # #####################################################################################################################
# # #####################################################################################################################
# # #####################################################################################################################
# # #####################################################################################################################
# # #####################################################################################################################

# # ##### FINISH UP ########################################
# # # OUTPUT EXTRA DATA FOR LAST SITE
# row_count = func.extra_record_output(ws_write, prev_sample_site_id, row_count)

# # ADD any no samples shots (but other shots in site had fish)
# # for sobs in func.sssoc_info:
# #     if sobs[1] not in shots_used and sobs[2] == '' and sobs[7] == '':
# #         row_count = func.extra_record_output_no_fish_shot(ws_write, sobs[0], sobs[1], row_count)
# for sobs in func.sssoc_info:
#     site_section_list = list(filter(lambda x: x[0] == sobs[0] and x[1] == int(sobs[1]), func.site_section_used))
#     # print('Obs data: Site {0} Shot {1} Sp: {2} Obs_ID: {3} site_used_len: {4}'.format(sobs[0], sobs[1], sobs[2], sobs[7], len(site_section_list)))
#     if len(site_section_list) == 0 and sobs[2] == '' and sobs[7] == '':
#         row_count = func.extra_record_output_no_fish_shot(ws_write, sobs[0], sobs[1], row_count)

# # ADD any no samples fish sites
# prev_site_id = ''
# for sobs in func.sssoc_info:
#     # print(sobs[0])
#     if sobs[0] != prev_site_id:
#         site_obs = list(filter(lambda x: x[0] == sobs[0], func.sssoc_info))
#         site_samples = list(filter(lambda x: x[16] == sobs[0], sample_list))
#         sample_flag = False
#         if len(site_samples) == 0:
#             row_count = func.extra_record_output(ws_write, sobs[0], row_count)
#             sample_flag = True

#     prev_site_id = sobs[0]


# func.sheet_sort_rows(ws_write, 2, 0, [47, 7, 18, 27, 32, 31])

# func.set_col_date_style(ws_write, (7-1))

# row_count = 1
# xl_header = list(("site_id", "section_number", "species", "collected", "observed", "collected_tally", "shot_id", "obs_id"))
# func.write_row(ws2_write, row_count, 1, xl_header)

# for i in func.sssoc_info:
#     xl_row = list((i[0], int(i[1]), i[2], i[3], i[4], i[5], i[6], i[7]))
#     row_count += 1
#     func.write_row(ws2_write, row_count, 1, xl_row)

# func.sheet_sort_rows(ws2_write, 2, 0, [1, 2, 3])
