import random
import re
import openpyxl
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import process_survey123_field_data_classes as cls
import os

global sssoc_info
global site_survey_info
global site_section_used
site_survey_info = []
sssoc_info = []  # site survey section observed collected
site_section_used = []


def read_in_excel_tab(wkbook_sheet):
    sheet = wkbook_sheet
    print('reading in {0}'.format(sheet.title))

    sheet_list = []
    i = 0

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=sheet.max_column, values_only=True):
        sheet_list.append(row)

    return sheet_list


def read_in_excel_tab_header(wkbook_sheet):
    sheet = wkbook_sheet
    for row in sheet.iter_rows(min_row=1, max_row=1, min_col=2, max_col=sheet.max_column, values_only=True):
        return row


gear_types = {
    "1E Large": "EF_LB",
    "2E Large": "EF_LB",
    "3E Medium": "EF_MB",
    "4E Small": "EF_SB",
    "5E Medium Grassal": "EFG_MB",
    "6E Medium": "EF_MB",
    "7E Polycraft": "EF",
    "15E Large Grassal": "EFG_LB",
    "Punt 6": "Punt 6",
    "Punt 7": "Punt 7",
    "V Nose 9": "V Nose 9",
    "GT 10": "GT 10",
    "V Nose 11": "V Nose 11",
    "Canoe": "Canoe",
    "10ft Tinny": "10ft Tinny",
    "12ft Tinny": "12ft Tinny",
    "Bank Mounted": "EF_BM",
    "Back Pack": "EF_BP",
    "EF_BP_LARV": "EF_BP_LARV",
    "Net": "Net",
    "Unknown": "Unknown",
    "EXTRA_SHOT_IN_SAMPLES": "EXTRA_SHOT_IN_SAMPLES"
}


def define_templates(sheetNames):


    if sheetNames[0].find('VEFMAP') >= 0 or sheetNames[0].find('Zeb') >= 0 or sheetNames[0].find('DEV_3_19') >= 0 or sheetNames[0].find('Dawson') >= 0:
        input_type = 'Fish_Survey_v2'

    elif sheetNames[0].find('Hack') >= 0 or sheetNames[0].find('Murray_Snags') >= 0:
        input_type = 'Fish_Survey_v2_1'

    elif sheetNames[0].find('Lieschke') >= 0:
        input_type = 'Fish_Survey_v2_2'

    elif sheetNames[0].find('Harris') >= 0:
        input_type = 'Fish_Survey_v2_3'

    elif sheetNames[0].find('Crowther') >= 0:
        input_type = 'Fish_Survey_v2_4'

    elif sheetNames[0].find('v1') >= 0:
        ##    input_type = 'Fish_Survey_v1' # Original Fish Survey format
        input_type = 'Fish_Survey_v1'

    elif sheetNames[0].find('ELH_IVT') >= 0:
        ##    input_type = 'Fish_Survey_v1' # Original Fish Survey format
        input_type = 'Fish_Survey_v2_5'
    else:
        input_type = ''


    input_type_msg = 'default' if input_type == '' else input_type
    print(colour_terminal_output('*** OUTPUT FORMAT: {0}'.format(input_type_msg), 'green'))

    # The following is to order how each page is presented in the results.

    # The data column will not be output if the index is == -1.
    # The data column will move to the corresponding index value, e.g. if 'site_code' is at index 1,
    # then [-1, 0, -1, -1, ...] will only output 'site_code' and place it at the start.
    # A 'j' will join that column and the following column, and place it in the position of the value after the 'j'.
    # A list of [0, 1, 2, 3, ...] will not change any order.
    # This will also not take into account 'ObjectID', i.e. 'GlobalID' is the first element.
    if input_type == 'Fish_Survey_v1':
        survey_template = [-1, 1, 4, 5, 'j', 6, 7, 8, 9, 10, 11, 12, 0, 13, -1, 14, 15, 16, 17, 18, 19, 20, 21, -1, -1, -1,
                           -1, -1, 2, 3]
        location_template = [-1, -1, -1, -1, -1, -1, -1, 0, 1, 2,
                             3]  # Keep in mind [... x, y] will become ... x_start, y_start, x_end, y_end]
        shot_template = [-1, 0, 1, 2, 3, 4, 5, 6, 7, 8, -1, -1, -1, -1, -1]
        obs_template = [-1, -1, -1, 0, 1, 2, -1, -1, -1, -1, -1, -1]
        sample_template = [-1, -1, -1, 0, 1, 2, 3, 4, 5, 6, 7, -1, 8, 9, 10, -1, -1, -1, -1, -1]

    elif input_type == 'Fish_Survey_v2': # VEFMAP, Zeb, Dawson

        survey_template = [-1, 1, 4, 5, 'j', 6, 7, 8, 9, 10, 11, 12, 0, 13, -1, 14, 15, 16, 17, 18, 19, 20, 21, -1, -1,
                           -1, -1, 2, 3]
        location_template = [-1, -1, -1, -1, -1, -1, -1, 0, 1, 2,
                             3]  # Keep in mind [... x, y] will become ... x_start, y_start, x_end, y_end]
        shot_template = [-1, 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, -1, -1, -1, -1, -1]
        obs_template = [-1, -1, -1, 0, 1, 2, -1, -1, -1, -1, -1, -1]
        sample_template = [-1, -1, -1, 0, 1, 2, 3, 4, 5, 6, 7, -1, 8, 9, 10, -1, -1, -1, -1, -1]

    elif input_type == 'Fish_Survey_v2_1': # Hack, Murray_Snags

        survey_template = [-1, 1, 4, 5, 'j', 6, 7, 8, 9, 10, 11, 12, 0, 13, -1, 14, 15, 16, 17, 18, 19, 20, 21, -1, -1,
                           -1, -1, -1, 2, 3]
        location_template = [-1, -1, -1, -1, -1, -1, -1, 0, 1, 2,
                             3]  # Keep in mind [... x, y] will become ... x_start, y_start, x_end, y_end]
##        shot_template = [-1, 0, 1, 2, 3, 4, 5, 6, 7, 8, -1, -1, -1, -1, -1, -1, 9, 10, 11, 12, 13]
        shot_template = [-1, 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, -1, -1, -1, -1, -1]
        obs_template = [-1, -1, -1, 0, 1, 2, -1, -1, -1, -1, -1, -1]
        sample_template = [-1, -1, -1, 0, 1, 2, 3, 4, 5, 6, 7, -1, 8, 9, 10, -1, -1, -1, -1, -1]

    elif input_type == 'Fish_Survey_v2_2': # Lieschke

        survey_template = [-1, 1, 4, 5, 'j', 6, 7, 8, 9, 10, 11, 12, 0, 13, -1, 14, 15, 16, 17, 18, 19, 20, 21, -1, -1,
                           -1, -1, 2, 3]
        location_template = [-1, -1, -1, -1, -1, -1, -1, 0, 1, 2,
                             3]  # Keep in mind [... x, y] will become ... x_start, y_start, x_end, y_end]
        shot_template = [-1, 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, -1, -1, -1, -1, -1]
        obs_template = [-1, -1, -1, 0, 1, 2, -1, -1, -1, -1, -1, -1]
        sample_template = [-1, -1, -1, 0, 1, 2, 3, 4, 5, 6, 7, -1, 8, 9, 10, -1, -1, -1, -1, -1]

    elif input_type == 'Fish_Survey_v2_3': # Harris

        survey_template = [-1, 1, 4, 5, 'j', 6, 7, 8, 9, 10, 11, 12, 0, 13, -1, 14, 15, 16, 17, 18, 19, 20, 21, -1, -1,
                           -1, -1, -1, 2, 3]
        location_template = [-1, -1, -1, -1, -1, -1, -1, 0, 1, 2,
                             3]  # Keep in mind [... x, y] will become ... x_start, y_start, x_end, y_end]
        shot_template = [-1,0,1,2,3,4,5,6,7,8,-1,-1,-1,-1,-1,9,10,11,12,13]
        obs_template = [-1, -1, -1, 0, 1, 2, -1, -1, -1, -1, -1, -1]
        sample_template = [-1, -1, -1, 0, 1, 2, 3, 4, 5, 6, 7, -1, 8, 9, 10, -1, -1, -1, -1, -1]

    elif input_type == 'Fish_Survey_v2_4': # Crowther

        survey_template = [-1, 1, 4, 5, 'j', 6, 7, 8, 9, 10, 11, 12, 0, 13, -1, 14, 15, 16, 17, 18, 19, 20, 21, -1, -1,
                           -1, -1, 2, 3]
        location_template = [-1, -1, -1, -1, -1, -1, -1, 0, 1, 2,
                             3]  # Keep in mind [... x, y] will become ... x_start, y_start, x_end, y_end]
        shot_template = [-1,0,1,2,3,4,5,6,7,8,9,10,11,12,13,-1,-1,-1,-1,-1,14]
        obs_template = [-1, -1, -1, 0, 1, 2, -1, -1, -1, -1, -1, -1]
        sample_template = [-1, -1, -1, 0, 1, 2, 3, 4, 5, 6, 7, -1, 8, 9, 10, -1, -1, -1, -1, -1]

    elif input_type == 'Fish_Survey_v2_5': # ELH IVT

        survey_template = [-1,1,4,14,5, 'j',6,7,8,9,10,11,12,0,13,-1,15,16,17,18,19,20,21,22,-1,-1,-1,-1,2,3]
        location_template = [-1, -1, -1, -1, -1, -1, -1, 0, 1, 2,
                             3]  # Keep in mind [... x, y] will become ... x_start, y_start, x_end, y_end]
        shot_template = [-1,0,-1,3,15,16,17,18,19,20,21,22,23,24,25,7,12,5,6,8,9,10,11,13,14,-1,-1,-1,-1,-1,4,1,2]
        obs_template = [-1, -1, -1, 0, 1, 2, -1, -1, -1, -1, -1, -1]
        sample_template = [-1, -1, -1, 0, 1, 2, 3, 4, 5, 6, 7, -1, 8, 9, 10, -1, -1, -1, -1, -1]
    else:

        survey_template = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26,
                           27, 28, 29]
        location_template = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9,
                             10]  # Keep in mind [... x, y] will become ... x_start, y_start, x_end, y_end]
        shot_template = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]
        obs_template = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]
        sample_template = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]



    return survey_template, location_template, shot_template, obs_template, sample_template, input_type




def get_random_shot(rs_site_id, rs_species, output, obs_header, shot_header):
    skip_next = 0

    # Filter completed data for shots with collected species at correct site and collected > 0:
    rs_sub_shots = list(filter(lambda x: x.shots[shot_header.index('ParentGlobalID')] == rs_site_id and x.observations[
        obs_header.index('species_obs')] == rs_species and x.observations[obs_header.index('section_collected')] > 0,
                               output))
    shotlist = []

    if rs_sub_shots is None or len(rs_sub_shots) == 0:
##        print('\n')
        print('Notice: No collected {0} available in shots for site {1}'.format(rs_species, rs_site_id))

    # If matches are found with Site, Species and Collected > 0:
    if len(rs_sub_shots) > 0:
        skip_next = 1
        prev_section_number = 0

        for rs_i in rs_sub_shots:
            if prev_section_number != rs_i.shots[shot_header.index('section_number')]:
                shotlist.append(rs_i)
                prev_section_number = rs_i.shots[shot_header.index('section_number')]

    # If matches found with only site and species:
    if skip_next == 0:
        rs_sub_shots = list(filter(
            lambda x: x.shots[shot_header.index('ParentGlobalID')] == rs_site_id and x.observations[
                obs_header.index('species_obs')] == rs_species, output))

        if len(rs_sub_shots) > 0:
            skip_next = 1
            prev_section_number = 0

            for rs_i in rs_sub_shots:
                if prev_section_number != rs_i.shots[shot_header.index('section_number')]:
                    shotlist.append(rs_i)
                    prev_section_number = rs_i.shots[shot_header.index('section_number')]

    # If only site match is found (excluding 'no fish' shots):
    if skip_next == 0:
        rs_sub_shots = list(filter(lambda x: x.shots[shot_header.index('ParentGlobalID')] == rs_site_id and x.observations[obs_header.index('species_obs')] != 'No Fish', output))

        if len(rs_sub_shots) > 0:
            skip_next = 1
            prev_section_number = 0
            for rs_i in rs_sub_shots:
                if prev_section_number != rs_i.shots[shot_header.index('section_number')]:
                    shotlist.append(rs_i)
                    prev_section_number = rs_i.shots[shot_header.index('section_number')]

    # If only site match is found XX(but only one shot)XX:
    if skip_next == 0:
        rs_sub_shots = list(filter(lambda x: x.shots[shot_header.index('ParentGlobalID')] == rs_site_id, output))

        if len(rs_sub_shots) > 0:
##            skip_next = 1
            prev_section_number = 0
            for rs_i in rs_sub_shots:
                if prev_section_number != rs_i.shots[shot_header.index('section_number')]:
                    shotlist.append(rs_i)
                    prev_section_number = rs_i.shots[shot_header.index('section_number')]
            print(colour_terminal_output('*** Caution: Any shot in site for {0} used: SiteID {1}\n***          Other valid shots' \
                ' for site may not be used if marked automatically for No Fish'.format(rs_species, rs_site_id), 'red'))

    if rs_sub_shots is None or len(shotlist) == 0:
        print(colour_terminal_output('*** ERROR (shot selector function): No {0} available: SiteID {1}'.format(rs_species, rs_site_id), 'red'))
        return False
    else:
        return random.choice(shotlist)


def adjust_species_count(current, raw_data, PGID, section_num, species, svy_header, obs_header, sample_header,
                         shot_header, tally_results, tally_header):

    #if no collected is set in samples
    collected_new = 1 if current[sample_header.index('collected')] in [0, None] else current[sample_header.index('collected')]

    for completed in raw_data:
        # Check that site, section and species match:
        if PGID == completed.surveys[svy_header.index('GlobalID')]:
            if section_num == completed.shots[shot_header.index('section_number')] or section_num == completed.samples[
                sample_header.index('section_number_samp')]:
                if species == completed.observations[obs_header.index('species_obs')]:
                    # Adjust accordingly
                    #reduced the obs section collected by sample collected value
                    completed.observations[obs_header.index('section_collected')] -= collected_new
                    break

        # Adjust Collected_Tally accordingly:
        # Find tally data with the same PGID, section_num and species:
    for completed in raw_data:
        for tally in tally_results:

            if tally[tally_header.index('Site_ID')] == PGID:
                if tally[tally_header.index('Section_Number')] == section_num:
                    if tally[tally_header.index('Species')] == species:

                        # Alter collected_tally:
                        tally[tally_header.index('Collected_Tally')] -= collected_new
                        return

    #if species at site and shot isn't in the data
    tally_results.append(
        [PGID, section_num, species, collected_new, 0, collected_new, section_num, None, None])

    return


def remove_unrequired_no_fish(raw_data, PGID, section_num, svy_header, obs_header, sample_header,
                         shot_header, tally_results, tally_header):
    for completed in raw_data:
        # Check that site, section and species match:
        if PGID == completed.surveys[svy_header.index('GlobalID')]:
            if section_num == completed.shots[shot_header.index('section_number')] or section_num == completed.samples[
                sample_header.index('section_number_samp')]:
                if 'No Fish' == completed.observations[obs_header.index('species_obs')]:
                    # Adjust accordingly
                    raw_data.remove(completed)

                    for item in tally_results:
                        if PGID == item[tally_header.index('Site_ID')]:
                            if section_num == item[tally_header.index('Section_Number')]:
                                if 'No Fish' == item[tally_header.index('Species')]:
                                    tally_results.remove(item)
                                    print('Notice: REMOVED UNREQUIRED NO FISH for site: {0} - shot: {1}'.format(PGID, section_num))
                                    return

    return

def populate_extra_collected(raw_data, raw_header):
  for rw in raw_data:
    if rw.collation[raw_header.index('section_collected')] > 0 and rw.collation[raw_header.index('collected')] is None:
        rw.collation[raw_header.index('collected')] = rw.collation[raw_header.index('section_collected')]
        # print('ID: {0} - s_coll: {1}, coll: {2}'.format(rw.collation[raw_header.index('Obs_GlobalID')], rw.collation[raw_header.index('section_collected')], rw.collation[raw_header.index('collected')]))

  return

def correct_net_gear_type(raw_data, raw_header):
  for rw in raw_data:
    if rw.collation[raw_header.index('gear_type')].lower() in ['net', 'unknown']:
        if rw.collation[raw_header.index('net')].lower() != 'ef':
            print('## Notice: Gear type converted to Net type ###')
            rw.collation[raw_header.index('gear_type')] = rw.collation[raw_header.index('net')]
        else:
            print(colour_terminal_output('*** ERROR Incorrect net type selected for shot id: {0}'.format(rw.collation[raw_header.index('Shot_GlobalID')]), 'red'))
    else:
        if rw.collation[raw_header.index('net')].lower() != 'ef':
            print('## Notice: Gear type converted to Net type ###')
            rw.collation[raw_header.index('gear_type')] = rw.collation[raw_header.index('net')]
  return

def append_holder_sample_row(shot_current, loc_current, survey_current, species, raw_data, svy_header, loc_header, shot_header, obs_header, sample_header):
    obs_current = [None] * len(obs_header)
    obs_current[obs_header.index('section_collected')] = -1
    obs_current[obs_header.index('observed')] = 0

    sample_current = [None] * len(sample_header)
    sample_current[sample_header.index('species_samp')] = species

    ID_Indices = [svy_header.index('GlobalID'),
                  loc_header.index('GlobalID'),
                  shot_header.index('GlobalID'),
                  obs_header.index('GlobalID'),
                  sample_header.index('GlobalID'), ]
    raw_data.append(cls.resultObject(survey_current,
                                     loc_current,
                                     shot_current,
                                     obs_current,
                                     sample_current,
                                     None,
                                     ID_Indices))
    return

def check_sample_in_raw_data(raw_data, sample_ID, sample_header):

    for completed in raw_data:
        if sample_ID == completed.samples[sample_header.index('GlobalID')]:
            return True

    return False

def add_samples_to_output_and_tally(samples_list, samples_header, section_number, shot_current, loc_current, survey_current, rawdata, survey_header, loc_header, shot_header, obs_header, tally_results):

    site_id = survey_current[survey_header.index('GlobalID')]
    shot_id = shot_current[shot_header.index('GlobalID')]

    samples_list = sorted(samples_list, key=lambda x: (samples_header.index('species_samp'), samples_header.index('species_samp_custom')))

    prev_sp = None
    collected = 0

    ##---------------------------------------------------------------------------------------------------------------
    ##-- Loop through samples and total up number collected. Then add to raw_data and tally. ------------------------
    ##---------------------------------------------------------------------------------------------------------------

    for smpl in samples_list:
        species = smpl[samples_header.index('species_samp')] if smpl[samples_header.index('species_samp_custom')] is None else smpl[samples_header.index('species_samp_custom')]

        if prev_sp != species and prev_sp != None:
            # Create filler for observations:
            append_holder_sample_row(shot_current, loc_current, survey_current, prev_sp, rawdata, survey_header
                , loc_header, shot_header, obs_header, samples_header)

            tally_results.append(
                [site_id, section_number, prev_sp, collected, 0, collected, shot_id, None, None])

            collected = 1 if smpl[samples_header.index('collected')] in [None, 0] else smpl[samples_header.index('collected')]

        else:
            collected += 1 if smpl[samples_header.index('collected')] in [None, 0] else smpl[samples_header.index('collected')]

        prev_sp = species


    # process the last relevant sample record from loop
    append_holder_sample_row(shot_current, loc_current, survey_current, prev_sp, rawdata, survey_header
        , loc_header, shot_header, obs_header, samples_header)

    tally_results.append(
        [site_id, section_number, prev_sp, collected, 0, collected, shot_id, None, None])

    ##---------------------------------------------------------------------------------------------------------------
    ##---------------------------------------------------------------------------------------------------------------
    return

def write_row(write_sheet, row_num: int, starting_column: str or int, write_values: list):
    if isinstance(starting_column, str):
        starting_column = ord(starting_column.lower()) - 96
    for wr_i, value in enumerate(write_values):
        write_sheet.cell(row_num, starting_column + wr_i, value)
    return

# Function to determine IDE for terminal colour formatting in colour_terminal_output()
def ide_eviron():
    if any('RSTUDIO' in name for name in os.environ) or any('VSCODE_INJECTION' in name for name in os.environ) \
            or any('PYCHARM_HOSTED' in name for name in os.environ):
        #pyscripter IDE running
        return 'non_pyscripter'
    else:
        return 'pyscripter'

def colour_terminal_output(message: str, colour):
    if ide_eviron() == 'non_pyscripter':
        if colour.lower() == 'red':
            message = '\033[31m' + message + '\033[0m'
            return message
        elif colour.lower() == 'green':
            message = '\033[32m' + message + '\033[0m'
            return message
        elif colour.lower() == 'yellow':
            message = '\033[33m' + message + '\033[0m'
            return message
        elif colour.lower() == 'blue':
            message = '\033[34m' + message + '\033[0m'
            return message
        else:
            return message
    else:
        return message

def sheet_sort_rows(ws, row_start, row_end=0, cols=None, sorter=None, reverse=False):
    # #""" Sorts given rows of the sheet
    # #    row_start   First row to be sorted
    # #    row_end     Last row to be sorted (default last row)
    # #    cols        Columns to be considered in sort
    # #    sorter      Function that accepts a tuple of values and
    # #                returns a sortable key
    # #    reverse     Reverse the sort order
    # #"""

    bottom = ws.max_row
    if row_end == 0:
        row_end = ws.max_row
    right = openpyxl.utils.get_column_letter(ws.max_column)
    if cols is None:
        cols = range(1, ws.max_column + 1)

    array = {}
    for ssr_row in range(row_start, row_end + 1):
        key = []
        for col in cols:
            key.append(ws.cell(ssr_row, col).value)
        key = tuple(key)
        array[key] = array.get(key, set()).union({ssr_row})

    order = sorted(array, key=sorter, reverse=reverse)

    ws.move_range(f"A{row_start}:{right}{row_end}", bottom)
    dest = row_start
    for src_key in order:
        for ssr_row in array[src_key]:
            src = ssr_row + bottom
            dist = dest - src
            ws.move_range(f"A{src}:{right}{src}", dist)
            dest += 1


def set_col_date_style(ws, col_index):
    # create date style:
    date_style = NamedStyle(name='date_style', number_format='YYYY-MM-DD')

    for cds_row in ws[2:ws.max_row]:  # skip the header
        cell = cds_row[col_index]  # column H
        cell.style = date_style


def extra_record_output(ws, ero_site_id, ero_row_count):
    # ######### OUTPUT ANY EXTRA FISH CAUGHT BUT NOT MEASURED ######################################
    ## x[5] is collected_left count
    sub_sssoc_info = list(filter(lambda x: x[0] == ero_site_id and x[5] > 0 and x[6] != 'IN SAMPLE INFO', sssoc_info))

    ero_row_count = write_extra_data(ws, sub_sssoc_info, ero_row_count, 'extra_caught')

    # ######### OUTPUT OBSERVED FISH ######################################
    ## x[4] is observed count
    sub_sssoc_info = list(filter(lambda x: x[0] == ero_site_id and x[4] > 0, sssoc_info))

    ero_row_count = write_extra_data(ws, sub_sssoc_info, ero_row_count, 'observed')

    # ######### OUTPUT NO FISH ######################################
    ## x[2] is species name
    # #                if prev_sample_site_id == 'becd3e03-1cd0-44cc-8f3b-69cc65ef1957':
    # print('got here {0}'.format(ero_site_id))
    sub_sssoc_info = list(filter(lambda x: x[0] == ero_site_id and x[2] == 'No Fish', sssoc_info))
    # print(len(sub_sssoc_info))
    # #                if prev_sample_site_id == 'becd3e03-1cd0-44cc-8f3b-69cc65ef1957' and len(sub_sssoc_info) <= 0:
    # #                    print('no hit for shot 8')
    ero_row_count = write_extra_data(ws, sub_sssoc_info, ero_row_count, 'no_fish')

    return ero_row_count


def extra_record_output_no_fish_shot(ws, ero_site_id, ero_section_number, ero_row_count):
    sub_sssoc_info = list(filter(lambda x: x[0] == ero_site_id and x[1] == str(ero_section_number), sssoc_info))
    ero_row_count = write_extra_data(ws, sub_sssoc_info, ero_row_count, 'no_shot_fish')

    return ero_row_count


def write_extra_data(ws_out, wed_sub_sssoc_info, r_count, extraDataType):
    if len(wed_sub_sssoc_info) > 0:
        for s in wed_sub_sssoc_info:

            wed_shot_i = s[1]
            if isinstance(wed_shot_i, str):
                # #                            print('converting shot')
                wed_shot_i = int(wed_shot_i)

            wed_sub_site_survey_info = list(
                filter(lambda x: x['k_site_id'] == s[0] and x['k_section_number'] == str(s[1]), site_survey_info))
            # print(s[0], s[1])
            if extraDataType == 'extra_caught':
                if len(wed_sub_site_survey_info) > 1:
                    print('*** MULTIPLE SURVEY ERROR GETTING EXTRA CAUGHT for site: {0} shot: {1}'.format(s[0], s[1]))

            elif extraDataType == 'observed':
                if len(wed_sub_site_survey_info) > 1:
                    print('*** MULTIPLE SURVEY ERROR GETTING OBSERVED for site: {0} shot: {1}'.format(s[0], s[1]))

            elif extraDataType == 'no_fish':
                if len(wed_sub_site_survey_info) > 1:
                    print('*** MULTIPLE SURVEY ERROR GETTING NO FISH for site: {0} shot: {1}'.format(s[0], s[1]))

            elif extraDataType == 'no_shot_fish':
                if len(wed_sub_site_survey_info) > 1:
                    print('*** MULTIPLE SURVEY ERROR GETTING NO SHOT FISH for site: {0} shot: {1}'.format(s[0], s[1]))

                elif len(wed_sub_site_survey_info) == 0:
                    wed_sub_site_survey_info = list(
                        filter(lambda x: x['k_site_id'] == s[0] and x['k_section_number'] == '1', site_survey_info))
                    if len(wed_sub_site_survey_info) > 0:
                        # wed_sub_site_survey_info[0]
                        write_excel_row(ws_out, r_count, wed_sub_site_survey_info[0], int(s[1]), 'No Fish', '', '', '',
                                        0, 0, '', '', '', '', '', '', s[7], '')
                        print('*** ADDED EXTRA SHOT WITH NO FISH for site: {0} shot: {1}'.format(s[0], s[1]))
                        r_count += 1
                        return r_count
                    else:
                        print('*** NO SHOT 1 SURVEY INFO ERROR for site: {0} shot: 1'.format(s[0]))

            for wed_ss_row in wed_sub_site_survey_info:

                # #    0: self.site_id,
                # #    1: self.section_number,
                # #    2: self.species,
                # #    3: self.collected,
                # #    4: self.observed,
                # #    5: self.collected_left,
                # #    6: self.shot_id
                # #    7: self.obs_id
                if extraDataType == 'extra_caught':

                    write_excel_row(ws_out, r_count, wed_ss_row, wed_shot_i, s[2], '', '', '', s[5], 0, '', '', '', '',
                                    '', '', s[7], '')
                    print('*** ADDED EXTRA CAUGHT for site: {0} shot: {1} species: {2}'.format(s[0], s[1], s[2]))

                elif extraDataType == 'observed':

                    write_excel_row(ws_out, r_count, wed_ss_row, wed_shot_i, s[2], '', '', '', 0, s[4], '', '', '', '',
                                    '', '', s[7], '')
                    print('Notice: Added OBSERVED for site: {0} shot: {1} species: {2}'.format(s[0], s[1], s[2]))

                elif extraDataType == 'no_fish':

                    write_excel_row(ws_out, r_count, wed_ss_row, wed_shot_i, s[2], '', '', '', 0, 0, '', '', '', '', '',
                                    '', s[7], '')
                    print('*** ADDED NO FISH for site: {0} shot: {1}'.format(s[0], s[1]))

                elif extraDataType == 'no_shot_fish':

                    write_excel_row(ws_out, r_count, wed_ss_row, wed_shot_i, 'No Fish', '', '', '', 0, 0, '', '', '',
                                    '', '', '', s[7], '')
                    print('*** ADDED NO FISH SHOT for site: {0} shot: {1}'.format(s[0], s[1]))

                r_count += 1
    return r_count


def write_excel_row(wsheet, rowcount, data_row, shot_num, wer_species, wer_fl, wer_tl, wer_w, wer_coll, wer_obs,
                    wer_recapture, wer_pit, wer_external_tag_no, wer_genetics_label, wer_otoliths_label,
                    wer_fauna_notes, wer_obst_id, wer_sample_id):
    if data_row['k_section_condition'].lower() == 'yes':
        section_condition_xl = 'FISHABLE'
    else:
        section_condition_xl = 'UNFISHABLE'

    personnel1 = data_row['k_personnel1']
    personnel2 = data_row['k_personnel2']

    # remove the common name within brackets
    wer_species = re.sub(r'\(.*?\) *', '', wer_species)
    wer_species = wer_species.strip()

    wer_gear_type = gear_types[data_row.gear_type]

    ##    wer_gear_type = data_row.gear_type #gear_types[data_row['k_gear_type']] if data_row['k_gear_type'] == data_row.gear_type else data_row.gear_type

    if data_row['k_survey_notes'] is None:
        wer_survey_notes = 'gear: {0}'.format(data_row['k_gear_type'])
    else:
        wer_survey_notes = '{0}, gear: {1}'.format(data_row['k_survey_notes'], data_row['k_gear_type'])

    wer_xl_row = list((
                      data_row['k_project_name'], data_row['k_site_code'], data_row['k_x_start'], data_row['k_y_start'],
                      data_row['k_x_finish'], data_row['k_y_finish'], data_row['k_survey_date'], wer_gear_type,
                      personnel1, personnel2, data_row['k_depth_secchi'], data_row['k_depth_max'],
                      data_row['k_depth_avg'], section_condition_xl, data_row['k_time_start'], data_row['k_time_end'],
                      wer_survey_notes, shot_num, data_row['k_electro_seconds'], data_row['k_soak_minutes_per_unit'],
                      data_row.section_time_start, data_row.section_time_end, data_row.volts, data_row.amps,
                      data_row.pulses_per_second, data_row.percent_duty_cycle, wer_species, wer_fl, wer_tl, wer_w,
                      wer_coll, wer_obs, wer_recapture, wer_pit, wer_external_tag_no, wer_genetics_label,
                      wer_otoliths_label, wer_fauna_notes, data_row['k_water_qual_depth'], data_row['k_ec_25c'],
                      data_row['k_water_temp'], data_row['k_do_mgl'], data_row['k_do_perc'], data_row['k_ph'],
                      data_row['k_turbidity_ntu'], data_row['k_chlorophyll'], data_row['k_site_id'],
                      data_row['k_shot_id'], wer_obst_id, wer_sample_id, data_row['k_data_x'], data_row['k_data_y']))

    write_row(wsheet, rowcount, 1, wer_xl_row)

